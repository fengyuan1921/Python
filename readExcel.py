#!/usr/bin/env python
import pandas as pd
file = 'E:\PycharmProjects\dtb.xlsx'
x1 = pd.ExcelFile(file)
print(x1.sheet_names)
df1 = x1.parse('主机申请')
print(df1)

from openpyxl import load_workbook
wb = load_workbook('E:\PycharmProjects\dtb.xlsx')
print(wb.sheetnames)
#sheet = wb.get_sheet_by_name('Sheet2')